import time, requests, json, os, io
from github import Github
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

def upload_to_github(data, filename="data_scm.json"):
    # Menggunakan token otomatis dari GitHub Actions
    g = Github(os.environ.get('GITHUB_TOKEN'))
    repo = g.get_repo("ipanrifan-create/DATA-SPK")
    file_content = json.dumps(data, indent=4)
    
    try:
        contents = repo.get_contents(filename)
        repo.update_file(contents.path, "Update data SCM otomatis (1 jam)", file_content, contents.sha)
        print(f"File {filename} berhasil diperbarui.")
    except:
        repo.create_file(filename, "Commit data SCM pertama", file_content)
        print(f"File {filename} berhasil dibuat.")

def jalankan_bot():
    print("=== [START] Operasi Bot SCM ===")
    options = uc.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    
    driver = uc.Chrome(options=options)
    try:
        driver.get("https://scm.nusadaya.net/login")
        wait = WebDriverWait(driver, 20)
        
        # Login
        email_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='text' or @placeholder='Email atau NIP']")))
        email_input.send_keys(os.environ.get('EMAIL_SCM'))
        driver.find_element(By.XPATH, "//input[@type='password']").send_keys(os.environ.get('PASS_SCM'))
        driver.find_element(By.XPATH, "//button[contains(text(), 'Log in')]").click()
        time.sleep(10)
        
        # Download
        session_cookies = {c['name']: c['value'] for c in driver.get_cookies()}
        response_dl = requests.get("https://scm.nusadaya.net/izin-prinsip/export", cookies=session_cookies)
        
        if response_dl.status_code == 200:
            wb = load_workbook(filename=io.BytesIO(response_dl.content), data_only=True)
            all_data = {sheet: [[cell if cell is not None else "" for cell in row] for row in wb[sheet].iter_rows(values_only=True)] for sheet in wb.sheetnames}
            upload_to_github(all_data)
        
    finally:
        driver.quit()
        print("=== [FINISH] Operasi Selesai ===")

if __name__ == "__main__":
    jalankan_bot()
